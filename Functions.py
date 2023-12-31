import os
import atexit
from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QMessageBox, QComboBox, QFileDialog, QLineEdit
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtCore import pyqtSlot, QFile, QTextStream, QDate
from PyQt5 import QtWidgets, QtGui, QtCore
from openpyxl import load_workbook, Workbook
from BDPS_ui import Ui_MainWindow
from BDPS_db.BDPS_queries import DBQueries

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5 import QtCore
from PyQt5 import QtGui
import time
import sys
import os
from openpyxl import load_workbook, Workbook

class BtnFunctions(QMainWindow):

    def __init__(self):
        super(BtnFunctions, self).__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)        
        self.ui.icon_only_widget.hide()
        self.ui.stackedWidget.setCurrentIndex(0)
        self.ui.home_btn_2.setChecked(True)
        
        self.get_recent_transactions()

        #========================== INITIALIZATIONS =====================================#
        self.ui.no_category_found.hide()
        self.ui.no_pricelist_found.hide()
        self.ui.no_dailytxn_found.hide()
        self.ui.no_datewiseT_found.hide()
        self.ui.no_datewiseP_found.hide()
        self.ui.no_transaction_found.hide() 
        self.ui.dateEdit_daily_tnx.setDate(QDate.currentDate())  
        self.ui.date_month_tnx.setDate(QDate.currentDate())
        self.ui.date_year_tnx.setDate(QDate.currentDate())
        self.ui.date_month_pmt.setDate(QDate.currentDate())
        self.ui.date_year_tnx_pmt.setDate(QDate.currentDate())
        self.ui.dateEdit.setDate(QDate.currentDate())
        self.ui.dateEdit_2.setDate(QDate.currentDate())
        #========================== RESET OF TABLES =====================================#
        self.ui.pushButton_4.clicked.connect(self.reset_dailytxn_table)
        self.ui.pushButton_5.clicked.connect(self.reset_txn_table)
        self.ui.pushButton_6.clicked.connect(self.reset_datewise_txn_table)
        self.ui.pushButton_7.clicked.connect(self.reset_datewise_payment_table)

        #========================== DASHBOARD =====================================#
        self.ui.transaction_record_tbl.model().rowsRemoved.connect(self.count_transaction_record)
        self.ui.transaction_record_tbl.model().rowsInserted.connect(self.count_transaction_record)
        self.ui.transaction_record_tbl.itemChanged.connect(self.count_transaction_record)
        self.ui.pricelist_table.model().rowsRemoved.connect(self.count_total_product)
        self.ui.pricelist_table.model().rowsInserted.connect(self.count_total_product)
        self.ui.pricelist_table.itemChanged.connect(self.count_total_product)
        self.ui.category_table.model().rowsRemoved.connect(self.count_total_service)
        self.ui.category_table.model().rowsRemoved.connect(self.count_total_service)
        self.ui.category_table.itemChanged.connect(self.count_total_service)

        self.count_transaction_record()
        self.count_total_product()
        self.count_total_service()        
        #========================== SEARCH FIELDS =====================================#
        self.ui.edit_search_pricelist.textChanged.connect(self.pricelist_table)
        self.ui.edit_search_category.textChanged.connect(self.category_table)
        self.ui.edit_search_daily_tnx.textChanged.connect(self.daily_txn_table)
        self.ui.edit_search_dwt.textChanged.connect(self.datewise_txn_table)
        self.ui.edit_search_dwp.textChanged.connect(self.datewise_payment_table)
        self.ui.edit_search_new_transaction.textChanged.connect(self.transaction_table)

        #========================== FILTERING =====================================#
        self.ui.dateEdit_daily_tnx.dateChanged.connect(self.filter_dailytxn_table)
        self.ui.date_year_tnx.dateChanged.connect(self.filter_bydate_dwtxn_table)
        self.ui.date_month_tnx.dateChanged.connect(self.filter_bydate_dwtxn_table)
        self.ui.date_year_tnx.dateChanged.connect(self.filter_bydate_dwtxn_table)
        self.ui.date_month_tnx.dateChanged.connect(self.filter_bydate_dwtxn_table)
        self.ui.date_year_tnx_pmt.dateChanged.connect(self.filter_date_wise_payment_table)
        self.ui. date_month_pmt.dateChanged.connect(self.filter_date_wise_payment_table)
        self.ui.dateEdit.dateChanged.connect(self.filter_txnrec_table)
        self.ui.dateEdit_2.dateChanged.connect(self.filter_txnrec_table)
        self.ui.transactionR_combobox.currentIndexChanged.connect(self.filter_txnrec_table)
        
        #========================== EXCEL EXPORT BUTTONS =====================================#
        self.ui.printreport_dwp_btn.clicked.connect(self.datewise_payment_toExcel)
        self.ui.printreport_daily_tnx_btn.clicked.connect(self.daily_transaction_toExcel)
        self.ui.printreport_dwt_btn.clicked.connect(self.datewise_transaction_toExcel)

        #================================== TABLES ===============================#
        self.ui.category_table.horizontalHeader().setVisible(True)
        self.ui.pricelist_table.horizontalHeader().setVisible(True)

        self.ui.category_table.clicked.connect(self.on_service_row_clicked)
        self.ui.pricelist_table.clicked.connect(self.on_pricelist_row_clicked)

        self.ui.category_table.setColumnWidth(0, 100)
        self.ui.category_table.setColumnWidth(2, 450)
        self.ui.category_table.setColumnWidth(3, 150)

        self.ui.pricelist_table.setColumnWidth(0, 150)
        self.ui.pricelist_table.setColumnWidth(1, 350)
        self.ui.pricelist_table.setColumnWidth(2, 250)
        self.ui.pricelist_table.setColumnWidth(3, 150)

        #========================== DATABASE PATH =====================================#
        dbFolder = os.path.abspath(os.path.join(os.path.dirname(__file__), 'BDPS_db/BDPS.db'))
        DBQueries.main(dbFolder)

        #======================== FETCH and MOD SERVICES =================================#
        DBQueries.displayServices(self, DBQueries.getAllServices(dbFolder))

        self.ui.add_category_btn.clicked.connect(lambda: DBQueries.addService(self, dbFolder))
        self.ui.edit_category_btn.clicked.connect(lambda: DBQueries.editService(self, dbFolder))
        self.ui.delete_category_btn.clicked.connect(lambda: DBQueries.deleteService(self, dbFolder))

        self.ui.category_table.itemSelectionChanged.connect(lambda: DBQueries.on_service_selection_changed(self))
        
        #======================== FETCH and MOD PRICELIST =================================#
        service_names = DBQueries.getServiceNames(dbFolder)
        self.ui.cat_name_pricelist.addItems(service_names)

        DBQueries.displayPrices(self, DBQueries.getAllPrices(dbFolder))

        self.ui.add_item_pricelist_btn.clicked.connect(lambda: DBQueries.addPrice(self, dbFolder))
        self.ui.update_pricelist_btn.clicked.connect(lambda: DBQueries.editPrice(self, dbFolder))
        self.ui.delete_pricelist_btn.clicked.connect(lambda: DBQueries.deletePrice(self, dbFolder))

        self.ui.pricelist_table.itemSelectionChanged.connect(lambda: DBQueries.on_price_selection_changed(self))

        #======================== FETCH and MOD JOBS =================================#
        self.ui.tnx_date_nt.setDate(QDate.currentDate())
        self.ui.category_name_nt.addItems(service_names)
        
        sizes = DBQueries.getProductSizes(self, dbFolder)

        self.ui.category_name_nt.currentIndexChanged.connect(lambda: DBQueries.getProductSizes(self, dbFolder))
        DBQueries.displayJobs(self, DBQueries.getAllJobs(dbFolder))
        self.ui.add_order_nt.clicked.connect(lambda: DBQueries.addJob(self, dbFolder))
        self.ui.delete_job_detail_btn.clicked.connect(lambda: DBQueries.deleteJob(self, dbFolder))
        self.ui.reset_job_detail_btn.clicked.connect(lambda: DBQueries.resetJobDetails(self, dbFolder))

        self.ui.order_detail_table.itemSelectionChanged.connect(lambda: DBQueries.on_job_selection_changed(self))

        #======================== FETCH and MOD PAYMENTS =================================#
        self.ui.discount_input.valueChanged.connect(self.update_discount_label)
        self.ui.discount_input.valueChanged.connect(lambda: DBQueries.checkDiscount(self, dbFolder))

        self.ui.subtotal_nt.textChanged.connect(lambda: DBQueries.update_total_nt(self, dbFolder))
        self.ui.discount_nt.textChanged.connect(lambda: DBQueries.update_total_nt(self, dbFolder))

        self.ui.payment_nt.textChanged.connect(lambda: DBQueries.check_payment_amount(self, dbFolder))
        self.ui.payment_nt.textChanged.connect(lambda: DBQueries.update_balance_nt(self, dbFolder))

        #======================== FETCH and MOD TRANSACTIONS =================================#
        self.ui.save_transaction_nt.clicked.connect(lambda: DBQueries.saveTransaction(self, dbFolder))
        DBQueries.displayTransactionRecords(self, DBQueries.getAllTransactions(dbFolder))
        DBQueries.displayDailyTransactions(self, DBQueries.getAllTransactions(dbFolder))
        DBQueries.displayDatewiseTransactions(self, DBQueries.getAllTransactions(dbFolder))
        DBQueries.displayDatewisePayments(self, DBQueries.getAllTransactions(dbFolder))

        current_txn_code = DBQueries.get_next_txn_code(self, dbFolder)
        self.ui.tnx_code_nt.setText(str(current_txn_code))

        self.ui.transaction_record_tbl.itemSelectionChanged.connect(lambda: DBQueries.on_txn_selection_changed(self))

        #=============================== ON EXIT ========================================#
        atexit.register(DBQueries.drop_job_temp_table, dbFolder)

        #========================== UPDATE TRANSACTIONS =====================================#
        self.ui.update_transaction.clicked.connect(lambda: self.update_transaction_pressed(dbFolder))
        self.ui.cancel_update.clicked.connect(self.on_cancel_update)
        self.ui.cancel_update.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(4))

        #======================== SEARCH FIELDS FUNCTIONS =================================#
        
    #Price list search field    
    def pricelist_table(self):
        search_pricelist = self.ui.edit_search_pricelist.text().strip()

        # Check if the search field is empty
        if not search_pricelist:
        # If the search field is empty, reset the QTableWidget to show all items
            for row in range(self.ui.pricelist_table.rowCount()):
                self.ui.pricelist_table.setRowHidden(row, False)
                self.ui.pricelist_table.show()
                self.ui.no_pricelist_found.hide()
        
        row_matches = False
            
        for row in range(self.ui.pricelist_table.rowCount()):
            for col in range(self.ui.pricelist_table.columnCount()):
                item = self.ui.pricelist_table.item(row, col)
                
                if item is not None and search_pricelist.lower() in item.text().lower():
                    self.ui.pricelist_table.setRowHidden(row, False)
                    row_matches = True
                    self.ui.pricelist_table.show()
                    self.ui.no_pricelist_found.hide()
                    break
                else:
                    self.ui.pricelist_table.setRowHidden(row, True)
        
        if not row_matches:
            self.ui.pricelist_table.hide()
            self.ui.no_pricelist_found.show()

    #Category search field   
    def category_table(self):
        search_category = self.ui.edit_search_category.text().strip()

        # Check if the search field is empty
        if not search_category:
        # If the search field is empty, reset the QTableWidget to show all items
            for row in range(self.ui.category_table.rowCount()):
                self.ui.category_table.setRowHidden(row, False)
                self.ui.category_table.show()
                self.ui.no_category_found.hide()            
        
        row_matches = False
            
        for row in range(self.ui.category_table.rowCount()):
            for col in range(self.ui.category_table.columnCount()):
                item = self.ui.category_table.item(row, col)
                
                if item is not None and search_category.lower() in item.text().lower():
                    self.ui.category_table.setRowHidden(row, False)
                    row_matches = True
                    self.ui.category_table.show()
                    self.ui.no_category_found.hide()
                    break
                else:
                    self.ui.category_table.setRowHidden(row, True)
        
        if not row_matches:
            self.ui.category_table.hide()
            self.ui.no_category_found.show()
            
    #Transaction records search field   
    def transaction_table(self):
        search_transaction = self.ui.edit_search_new_transaction.text().strip()

        # Check if the search field is empty
        if not search_transaction:
        # If the search field is empty, reset the QTableWidget to show all items
            for row in range(self.ui.transaction_record_tbl.rowCount()):
                self.ui.transaction_record_tbl.setRowHidden(row, False)
                self.ui.transaction_record_tbl.show()
                total_txn_record = self.ui.transaction_record_tbl.rowCount()
                self.ui.lineEdit.setText(f"{total_txn_record}")                          
                self.ui.no_transaction_found.hide()            
        else:
            self.filter_txnrec_table()
                                   
    #Daily transactions search field
    def daily_txn_table(self):
        search_daily_txn = self.ui.edit_search_daily_tnx.text().strip()

        # Check if the search field is empty
        if not search_daily_txn:
        # If the search field is empty, reset the QTableWidget to show all items
            for row in range(self.ui.daily_tnx_table.rowCount()):
                self.ui.daily_tnx_table.setRowHidden(row, False)
                self.ui.daily_tnx_table.show()
                total_dailytxn = self.ui.daily_tnx_table.rowCount()
                self.ui.dt_total.setText(f"{total_dailytxn}")
                self.ui.no_dailytxn_found.hide()
        else:
            self.filter_dailytxn_table()
                        
    #Date-wise transactions search field
    def datewise_txn_table(self):
        search_datewise_txn = self.ui.edit_search_dwt.text().strip()
        # Check if the search field is empty
        if not search_datewise_txn:
        # If the search field is empty, reset the QTableWidget to show all items
            for row in range(self.ui.datewise_transaction_table.rowCount()):
                self.ui.datewise_transaction_table.setRowHidden(row, False)
                self.ui.datewise_transaction_table.show()
                self.ui.no_datewiseT_found.hide()
                total_datewise_txn = self.ui.datewise_transaction_table.rowCount()
                self.ui.lineEdit_2.setText(f"{total_datewise_txn}")
        else:
            self.filter_bydate_dwtxn_table()
            
    #Date-wise payments search field
    def datewise_payment_table(self):
        search_datewise_pay = self.ui.edit_search_dwp.text().strip()
        total_sales_dwp = 0
        # Check if the search field is empty
        if not search_datewise_pay:
        # If the search field is empty, reset the QTableWidget to show all items
            for row in range(self.ui.datewise_payment_table.rowCount()):
                amount_paid = self.ui.datewise_payment_table.item(row, 5)
                self.ui.datewise_payment_table.setRowHidden(row, False)
                self.ui.datewise_payment_table.show()
                self.ui.no_datewiseP_found.hide()
                amount_paid = float(amount_paid.text())
                total_sales_dwp += amount_paid
            self.ui.label_49.setText(f"{total_sales_dwp:.2f}")
        else:
            self.filter_date_wise_payment_table()

        #======================== FILTER FUNCTIONS =================================#
              
    #Daily Transaction Filter 
    def filter_dailytxn_table(self):
        search_daily_txn = self.ui.edit_search_daily_tnx.text().strip().lower()
        selected_date = self.ui.dateEdit_daily_tnx.date()
        new_date = selected_date.toString("yyyy-MM-dd")
        total_dailytxn = 0
        no_row_matches = True
                    
        for row in range(self.ui.daily_tnx_table.rowCount()):
            date_item = self.ui.daily_tnx_table.item(row, 1)
            for col in range(self.ui.daily_tnx_table.columnCount()):
                item = self.ui.daily_tnx_table.item(row, col)
            
                if date_item is not None:
                    date = date_item.text().strip()
                    if  (search_daily_txn in item.text().lower()) and (date == new_date):
                        self.ui.daily_tnx_table.setRowHidden(row, False)
                        self.ui.daily_tnx_table.show()
                        total_dailytxn += 1
                        self.ui.no_dailytxn_found.hide()
                        no_row_matches = False
                        break 
                    else:
                        self.ui.daily_tnx_table.setRowHidden(row, True)
        
        if no_row_matches == True and not date == new_date:
            self.ui.daily_tnx_table.hide()
            date_in_words = QDate.fromString(new_date, "yyyy-MM-dd").toString("yyyy MMMM d")
            self.ui.no_dailytxn_found.setText(f"There is no transaction found during {date_in_words}.")
            self.ui.no_dailytxn_found.show()
            
        if no_row_matches == True and date == new_date:
            self.ui.daily_tnx_table.hide()
            date_in_words = QDate.fromString(new_date, "yyyy-MM-dd").toString("yyyy MMMM d")
            self.ui.no_dailytxn_found.setText(f"There is no {search_daily_txn} found during {date_in_words}.")
            self.ui.no_dailytxn_found.show()
                
        self.ui.dt_total.setText(f"{total_dailytxn}")
    
    #Date-wise transaction filter 
    def filter_bydate_dwtxn_table(self):
        from_month = self.ui.date_month_tnx.date().month()
        to_year = self.ui.date_year_tnx.date().year()
        total_datewise_txn = 0
        search_datewise_txn = self.ui.edit_search_dwt.text().strip().lower()

        from_date = QDate(to_year, from_month, 1)
        to_date = QDate(to_year, from_month, from_date.daysInMonth())    
        
        no_row_matches = True
        
        for row in range(self.ui.datewise_transaction_table.rowCount()):  
            for col in range(self.ui.datewise_transaction_table.columnCount()):
                item = self.ui.datewise_transaction_table.item(row, col).text().lower()
                date_item = self.ui.datewise_transaction_table.item(row, 1)
                if date_item is not None:
                    date_str = date_item.text()
                    date = QDate.fromString(date_str, "yyyy-MM-dd")
                    if (from_date <= date <= to_date) and (search_datewise_txn in item):
                        self.ui.datewise_transaction_table.setRowHidden(row, False)
                        no_row_matches = False
                        total_datewise_txn += 1
                        self.ui.datewise_transaction_table.show()
                        self.ui.no_datewiseT_found.hide()  
                        break
                    else:
                        self.ui.datewise_transaction_table.setRowHidden(row, True)
                        
        if no_row_matches == True and not from_date <= date <= to_date:
            self.ui.datewise_transaction_table.hide()
            self.ui.no_datewiseT_found.setText(f"There is no transaction found during the month of {from_date.longMonthName(from_date.month())} {to_year}.")
            self.ui.no_datewiseT_found.show()
            
        if no_row_matches == True and from_date <= date <= to_date:
            self.ui.datewise_transaction_table.hide()
            self.ui.no_datewiseT_found.setText(f"There is no {search_datewise_txn} found during the month of {from_date.longMonthName(from_date.month())} {to_year}.")
            self.ui.no_datewiseT_found.show()
        
        self.ui.lineEdit_2.setText(f"{total_datewise_txn}")

    #Date-wise payment filter 
    def filter_date_wise_payment_table(self):
        from_month = self.ui.date_month_pmt.date().month()
        to_year = self.ui.date_year_tnx_pmt.date().year()
        total_sales_dwp = 0
        search_datewise_pay = self.ui.edit_search_dwp.text().strip().lower()

        from_date = QDate(to_year, from_month, 1)
        to_date = QDate(to_year, from_month, from_date.daysInMonth())    
        
        no_row_matches = True
        
        for row in range(self.ui.datewise_payment_table.rowCount()):  
            for col in range(self.ui.datewise_payment_table.columnCount()):
                item = self.ui.datewise_payment_table.item(row, col).text().lower()
                date_item = self.ui.datewise_payment_table.item(row, 1)
                amount_paid = self.ui.datewise_payment_table.item(row, 5)
                if date_item is not None:
                    date_str = date_item.text()
                    date = QDate.fromString(date_str, "yyyy-MM-dd")
                    if (from_date <= date <= to_date) and (search_datewise_pay in item):
                        self.ui.datewise_payment_table.setRowHidden(row, False)
                        no_row_matches = False
                        self.ui.datewise_payment_table.show()
                        self.ui.no_datewiseP_found.hide()  
                        amount_paid = float(amount_paid.text())
                        total_sales_dwp += amount_paid
                        break
                    else:
                        self.ui.datewise_payment_table.setRowHidden(row, True)
                        
        if no_row_matches == True and not from_date <= date <= to_date:
            self.ui.label_49.setText("0.00")
            self.ui.datewise_payment_table.hide()
            self.ui.no_datewiseP_found.setText(f"There is no transaction found during the month of {from_date.longMonthName(from_date.month())} {to_year}.")
            self.ui.no_datewiseP_found.show()
            
        if no_row_matches == True and from_date <= date <= to_date:
            self.ui.label_49.setText("0.00")
            self.ui.datewise_payment_table.hide()
            self.ui.no_datewiseP_found.setText(f"There is no {search_datewise_pay} found during the month of {from_date.longMonthName(from_date.month())} {to_year}.")
            self.ui.no_datewiseP_found.show()

        self.ui.label_49.setText(f"{total_sales_dwp:.2f}")
        
    #Transaction records filter 
    def filter_txnrec_table(self):
        selected_item = self.ui.transactionR_combobox.currentText()
        search_transaction = self.ui.edit_search_new_transaction.text().strip().lower()
        
        date_from = self.ui.dateEdit.date()
        date_to = self.ui.dateEdit_2.date()
        total_txn_record = 0  
        date_from_str = date_from.toString("yyyy-MM-dd")
        date_to_str = date_to.toString("yyyy-MM-dd")     
           
        no_row_matches = True
                    
        for row in range(self.ui.transaction_record_tbl.rowCount()):
            status_item = self.ui.transaction_record_tbl.item(row, 5)
            date_item = self.ui.transaction_record_tbl.item(row, 1)
            for col in range(self.ui.transaction_record_tbl.columnCount()):
                item = self.ui.transaction_record_tbl.item(row, col)
            
                if status_item is not None and date_item is not None:
                    status = status_item.text().strip()
                    str_date_item = date_item.text().strip()
                    if (selected_item == "All Transactions") and (search_transaction in item.text().lower()) and (date_from_str <= str_date_item <= date_to_str):
                        self.ui.transaction_record_tbl.setRowHidden(row, False)
                        self.ui.transaction_record_tbl.show()
                        self.ui.no_transaction_found.hide()
                        no_row_matches = False
                        total_txn_record += 1   
                        break
                    elif (selected_item == "Pending Transactions" and status == "Pending Transaction") and (search_transaction in item.text().lower()) and (date_from_str <= str_date_item <= date_to_str):
                        self.ui.transaction_record_tbl.setRowHidden(row, False)
                        self.ui.transaction_record_tbl.show()
                        self.ui.no_transaction_found.hide()
                        no_row_matches = False
                        total_txn_record += 1   
                        break
                    elif (selected_item == "Cancelled Transactions" and status == "Cancelled Transaction") and (search_transaction in item.text().lower()) and (date_from_str <= str_date_item <= date_to_str):
                        self.ui.transaction_record_tbl.setRowHidden(row, False)
                        self.ui.transaction_record_tbl.show()
                        self.ui.no_transaction_found.hide()
                        no_row_matches = False
                        total_txn_record += 1   
                        break
                    elif (selected_item == "Successful Transactions" and status == "Sucessful Transaction") and (search_transaction in item.text().lower()) and (date_from_str <= str_date_item <= date_to_str):
                        self.ui.transaction_record_tbl.setRowHidden(row, False)
                        self.ui.transaction_record_tbl.show()
                        self.ui.no_transaction_found.hide()
                        no_row_matches = False
                        total_txn_record += 1   
                        break 
                    else:
                        self.ui.transaction_record_tbl.setRowHidden(row, True)
                        
        if no_row_matches == True and not date_from_str <= str_date_item <= date_to_str:
            date_in_words1 = QDate.fromString(date_from_str, "yyyy-MM-dd").toString("yyyy MMMM d")
            date_in_words2 = QDate.fromString(date_to_str, "yyyy-MM-dd").toString("yyyy MMMM d")
            self.ui.transaction_record_tbl.hide()
            self.ui.no_transaction_found.setText(f"There are no {selected_item} found during {date_in_words1} to  {date_in_words2}.")
            self.ui.no_transaction_found.show()
            
        if no_row_matches == True and date_from_str <= str_date_item <= date_to_str:
            date_in_words1 = QDate.fromString(date_from_str, "yyyy-MM-dd").toString("yyyy MMMM d")
            date_in_words2 = QDate.fromString(date_to_str, "yyyy-MM-dd").toString("yyyy MMMM d")
            self.ui.transaction_record_tbl.hide()
            self.ui.no_transaction_found.setText(f"There are no search results found in {selected_item} during the {date_in_words1} to  {date_in_words2}.")
            self.ui.no_transaction_found.show()

        self.ui.lineEdit.setText(f"{total_txn_record}")                          
        #======================== EXPORT FUNCTIONS =================================#
        
    #Daily Transaction to Excel
    def daily_transaction_toExcel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")

        if path:
            sample = None
            try:
                sample = load_workbook(path)
            except FileNotFoundError:
                sample = Workbook()
                    
            data = sample.active
            
            headers = [self.ui.daily_tnx_table.horizontalHeaderItem(i).text() for i in range(self.ui.daily_tnx_table.columnCount())]
                
            for column_index, header in enumerate(headers, start = 1):
                data.cell(row=1, column = column_index, value = header)
                    
            for row in range(self.ui.daily_tnx_table.rowCount()):
                for col in range(self.ui.daily_tnx_table.columnCount()):
                    items = self.ui.daily_tnx_table.item(row, col)
                    if items is not None:
                        data.cell(row=row + 2, column = col + 1, value = items.text())
                
            sample.save(path)
            
    #Date-wise Transaction to Excel        
    def datewise_transaction_toExcel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")

        if path:
            sample = None
            try:
                sample = load_workbook(path)
            except FileNotFoundError:
                sample = Workbook()
                    
            data = sample.active
            
            headers = [self.ui.datewise_transaction_table.horizontalHeaderItem(i).text() for i in range(self.ui.datewise_transaction_table.columnCount())]
                
            for column_index, header in enumerate(headers, start = 1):
                data.cell(row=1, column = column_index, value = header)
                    
            for row in range(self.ui.datewise_transaction_table.rowCount()):
                for col in range(self.ui.datewise_transaction_table.columnCount()):
                    items = self.ui.datewise_transaction_table.item(row, col)
                    if items is not None:
                        data.cell(row=row + 2, column = col + 1, value = items.text())
                
            sample.save(path)                      
    
    #Date-wise Payment to Excel
    def datewise_payment_toExcel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")

        if path:
            sample = None
            try:
                sample = load_workbook(path)
            except FileNotFoundError:
                sample = Workbook()
                    
            data = sample.active
            
            headers = [self.ui.datewise_payment_table.horizontalHeaderItem(i).text() for i in range(self.ui.datewise_payment_table.columnCount())]
                
            for column_index, header in enumerate(headers, start = 1):
                data.cell(row=1, column = column_index, value = header)
                    
            for row in range(self.ui.datewise_payment_table.rowCount()):
                for col in range(self.ui.datewise_payment_table.columnCount()):
                    items = self.ui.datewise_payment_table.item(row, col)
                    if items is not None:
                        data.cell(row=row + 2, column = col + 1, value = items.text())
                
            sample.save(path) 
             
        #======================== COUNT LABEL TRANSACTIONS FUNCTIONS =================================#

    def count_labels_txns(self):
        
        total_dailytxn = 0
        total_datewise_txn = 0
        total_sales_dwp = 0
        total_txn_record = 0
        
        total_dailytxn = self.ui.daily_tnx_table.rowCount()
        self.ui.dt_total.setText(f"{total_dailytxn}")
        
        total_datewise_txn = self.ui.datewise_transaction_table.rowCount()
        self.ui.lineEdit_2.setText(f"{total_datewise_txn}")
                
        total_txn_record = self.ui.transaction_record_tbl.rowCount()
        self.ui.lineEdit.setText(f"{total_txn_record}")
        
        for row in range(self.ui.datewise_payment_table.rowCount()):
            amount_paid = self.ui.datewise_payment_table.item(row, 5)
            
            if amount_paid is not None:
                try:
                    amount_paid = float(amount_paid.text())
                    total_sales_dwp += amount_paid
                except ValueError:
                    pass
        
        self.ui.label_49.setText(f"{total_sales_dwp:.2f}")

        #======================== RESET FUNCTIONS =================================#
        
    def reset_dailytxn_table(self):
        self.count_labels_txns()
        self.ui.edit_search_daily_tnx.setText("")
        self.ui.dateEdit_daily_tnx.setDate(QDate.currentDate())  
        for row in range(self.ui.daily_tnx_table.rowCount()):
            self.ui.daily_tnx_table.setRowHidden(row, False)        
            self.ui.daily_tnx_table.show()
            self.ui.no_dailytxn_found.hide()
            
    def reset_txn_table(self):
        self.count_labels_txns()
        self.ui.dateEdit.setDate(QDate.currentDate())
        self.ui.dateEdit_2.setDate(QDate.currentDate())
        self.ui.edit_search_new_transaction.setText("")
        self.ui.transactionR_combobox.setCurrentIndex(0)
        for row in range(self.ui.transaction_record_tbl.rowCount()):
            self.ui.transaction_record_tbl.setRowHidden(row, False)        
            self.ui.transaction_record_tbl.show()
            self.ui.no_transaction_found.hide()            
    
    def reset_datewise_txn_table(self):
        self.count_labels_txns()
        self.ui.date_month_tnx.setDate(QDate.currentDate())
        self.ui.date_year_tnx.setDate(QDate.currentDate())
        self.ui.edit_search_dwt.setText("")
        for row in range(self.ui.datewise_transaction_table.rowCount()):
            self.ui.datewise_transaction_table.setRowHidden(row, False)        
            self.ui.datewise_transaction_table.show()
            self.ui.no_datewiseT_found.hide()
    
    def reset_datewise_payment_table(self):
        self.ui.edit_search_dwp.setText("")
        self.ui.date_month_pmt.setDate(QDate.currentDate())
        self.ui.date_year_tnx_pmt.setDate(QDate.currentDate())
        self.count_labels_txns()
        for row in range(self.ui.datewise_payment_table.rowCount()):
            self.ui.datewise_payment_table.setRowHidden(row, False)        
            self.ui.datewise_payment_table.show()
            self.ui.no_datewiseP_found.hide()
                        
        #======================== DASHBOARD FUNCTIONS =================================#
    def get_recent_transactions(self):
        current_date = QDate.currentDate()
        two_days_ago = current_date.addDays(-2)

        # Clear the recent_transaction_table before adding recent transactions
        self.ui.tnx_summary_table.setRowCount(0)

        for row in range(self.ui.transaction_record_tbl.rowCount()):
            date_item = self.ui.transaction_record_tbl.item(row, 1)
            if date_item:
                date_str = date_item.text()
                date = QDate.fromString(date_str, "yyyy-MM-dd")
                
                if two_days_ago <= date <= current_date:
                    # Copy the row to the recent_transaction_table
                    job = self.ui.transaction_record_tbl.item(row, 0)
                    particular = self.ui.transaction_record_tbl.item(row, 2)
                    total = self.ui.transaction_record_tbl.item(row, 3)

                    row_position = self.ui.tnx_summary_table.rowCount()
                    self.ui.tnx_summary_table.insertRow(row_position)
                    self.ui.tnx_summary_table.setItem(row_position, 0, QTableWidgetItem(job.text()))
                    self.ui.tnx_summary_table.setItem(row_position, 1, QTableWidgetItem(particular.text()))
                    self.ui.tnx_summary_table.setItem(row_position, 2, QTableWidgetItem(total.text()))
  
    def count_transaction_record(self):
        
        total_transaction = self.ui.transaction_record_tbl.rowCount()
        self.ui.total_transactions.setText(f"{total_transaction}")

        pending_txn = 0
        successful_txn = 0
        cancelled_txn = 0
        total_sales = 0

        for row in range(self.ui.transaction_record_tbl.rowCount()):            
            status_item = self.ui.transaction_record_tbl.item(row, 5)
            amount_paid = self.ui.transaction_record_tbl.item(row, 4)
            if status_item is not None and status_item.text().strip() == "Pending Transaction":
                pending_txn += 1
            elif status_item is not None and status_item.text().strip() == "Sucessful Transaction":
                successful_txn += 1
            elif status_item is not None and status_item.text().strip() == "Cancelled Transaction":
                cancelled_txn += 1
        
        for row in range(self.ui.transaction_record_tbl.rowCount()):
            amount_paid = self.ui.transaction_record_tbl.item(row, 4)
            
            if amount_paid is not None:
                try:
                    amount_paid = float(amount_paid.text())
                    total_sales += amount_paid
                except ValueError:
                    pass
        
        self.ui.total_sales.setText(f"{total_sales:.2f}")
                
        self.ui.PendingTransactions.setText(f"{pending_txn}")
        self.ui.sucessful_transactions.setText(f"{successful_txn}")        
        self.ui.cancelled_transactions.setText(f"{cancelled_txn}")
        self.ui.total_sales.setText(f"{total_sales:.2f}")
    """
    def count_transaction_status(self):
        pending_txn = 0
        successful_txn = 0
        cancelled_txn = 0
        total_sales = 0

        for row in range(self.ui.transaction_record_tbl.rowCount()):
            status_item = self.ui.transaction_record_tbl.item(row, 5)
            if status_item is not None and status_item.text().strip() == "Pending":
                pending_txn += 1
            elif status_item is not None and status_item.text().strip() == "Successful":
                successful_txn += 1
            elif status_item is not None and status_item.text().strip() == "Cancelled":
                cancelled_txn += 1

        for row in range(self.ui.transaction_record_tbl.rowCount()):
            amount_paid = self.ui.transaction_record_tbl.item(row, 4)
            
            if amount_paid is not None:
                try:
                    amount_paid = float(amount_paid.text())
                    total_sales += amount_paid
                except ValueError:
                    pass
        
        self.ui.total_sales.setText(f"{total_sales:.2f}")
        self.ui.PendingTransactions.setText(f"{pending_txn}")  
        self.ui.sucessful_transactions.setText(f"{successful_txn}")        
        self.ui.cancelled_transactions.setText(f"{cancelled_txn}")
        """
    def count_total_product(self):
        total_prod = 0
        
        total_prod = self.ui.pricelist_table.rowCount()
        self.ui.total_product.setText(f"{total_prod}")
    
    def count_total_service(self):
        total_srvce = 0
        
        total_srvce = self.ui.category_table.rowCount()
        self.ui.total_category.setText(f"{total_srvce}")

    #----------------------------------------------------------------------#
    
    ## Function for searching
    def on_search_btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(5)
        search_text = self.ui.search_input.text().strip()
        if search_text:
            self.ui.label_9.setText(search_text)

    ## Function for changing page to user page
    def on_user_btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(6)

    ## Change QPushButton Checkable status when stackedWidget index changed
    def on_stackedWidget_currentChanged(self, index):
        btn_list = self.ui.icon_only_widget.findChildren(QPushButton) \
                    + self.ui.full_menu_widget.findChildren(QPushButton)
        
        for btn in btn_list:
            if index in [7, 8]:
                btn.setAutoExclusive(False)
                btn.setChecked(False)
            else:
                btn.setAutoExclusive(True)
            
    ## ----------------------- functions for changing menu page ---------------------
    def on_home_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(0)
        self.count_labels_txns()
        self.get_recent_transactions()
    def on_home_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(0)
        self.count_labels_txns()
        self.get_recent_transactions()
    def on_price_list_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(1)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_price_list_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(1)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_categories_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(2)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_categories_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(2)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_new_transaction_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(3)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_new_transaction_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(3)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_transaction_record_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(4)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_transaction_record_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(4)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_daily_transaction_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(5)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_daily_transaction_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(5)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_daily_transaction_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(6)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_datewise_transaction_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(6)
        self.count_labels_txns()
        self.get_recent_transactions()
    
    def on_datewise_payment_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(7)
        self.count_labels_txns()
        self.get_recent_transactions()

    def on_daily_payment_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(7)
        self.count_labels_txns()
        self.get_recent_transactions()

    def update_transaction_pressed(self, dbFolder):
        selected_row = self.ui.transaction_record_tbl.currentRow()

        if selected_row < 0:
            QMessageBox.about(self, "Message", "No transaction selected.")
        else:
            DBQueries.updateTransactions(self, dbFolder)
    
     ## ----------------------- ADD SHADOW --------------------- 


    #========================== SETTING PLACEHOLDERS =========================#    
    def on_service_row_clicked(self, index):
        selected_rows = self.ui.category_table.selectionModel().selectedRows()
        row = index.row()
        serv_id = self.ui.category_table.item(row, 0).text()
        serv_name = self.ui.category_table.item(row, 1).text()
        serv_desc = self.ui.category_table.item(row, 2).text()
        serv_sts = self.ui.category_table.item(row, 3).text()

        self.ui.id_category.setReadOnly(True)
        self.ui.id_category.setText(serv_id)
        self.ui.product_name_category.clear()
        self.ui.product_name_category.setPlaceholderText(serv_name)
        self.ui.category_description.clear()
        self.ui.category_description.setPlaceholderText(serv_desc)
        self.ui.status_category.setCurrentIndex(self.ui.status_category.findText(serv_sts))

        if len(selected_rows) != 1:
            self.ui.id_category.setReadOnly(True)
            self.ui.id_category.setText('')
            self.ui.product_name_category.clear()
            self.ui.product_name_category.setPlaceholderText('')
            self.ui.category_description.clear()
            self.ui.category_description.setPlaceholderText('')
            self.ui.status_category.setCurrentIndex(-1)
    
    def on_pricelist_row_clicked(self, index):
        selected_rows = self.ui.pricelist_table.selectionModel().selectedRows()
        row = index.row()
        prod_id = self.ui.pricelist_table.item(row, 0).text()
        cat_name = self.ui.pricelist_table.item(row, 1).text()
        prod_sz = self.ui.pricelist_table.item(row, 2).text()
        prod_price = self.ui.pricelist_table.item(row, 3).text()

        self.ui.id_pricelist.setReadOnly(True)
        self.ui.id_pricelist.setText(prod_id)
        self.ui.cat_name_pricelist.setCurrentIndex(self.ui.cat_name_pricelist.findText(cat_name))
        self.ui.size_pricelist.setPlaceholderText(prod_sz)
        self.ui.price_pricelist.setSpecialValueText(prod_price)

        if len(selected_rows) != 1:
            self.ui.id_pricelist.setText('')
            self.ui.cat_name_pricelist.setCurrentIndex(-1)
            self.ui.size_pricelist.clear()
            self.ui.size_pricelist.setPlaceholderText('')
            self.ui.price_pricelist.setSpecialValueText('')
    
    def update_discount_label(self, discount):
        self.ui.discount_nt.setText(str(discount))
    
    def on_cancel_update(self):
        self.ui.order_detail_table_2.clearContents()
        self.ui.order_detail_table_2.setRowCount(0)
        self.ui.order_detail_table_3.clearContents()
        self.ui.order_detail_table_3.setRowCount(0)

